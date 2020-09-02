import xlrd
import  xlwt
import openpyxl
import xlsxwriter    #插入图片图表的库，根据excel表格内的数据生成图表


'''
xlrd库用来读取excel的内容，方法时先open_workbook 传入file对象
'''

file=r'患者注册绑定.xlsx'
f=xlrd.open_workbook(filename=file)
sheet1=f.sheet_by_index(0)
print("打印sheet名字",sheet1.name,"sheet获取单元格内容，cell方法传的是索引下标",sheet1.cell(2,2).value)
print("打印sheet的总行数和总列数,sheet的下标",sheet1.nrows,sheet1.ncols,sheet1.number)

li=sheet1.row(0)
print("打印打印一行数据，返回的是单元格类型：‘数据’ ，单元格类型又 empty、number、text")
print(li)
print("打印全部数据时，返回的是一个生成器对象")
print(sheet1.get_rows())
print("excle内的方法还支持类似切片操作，指定行或者列 进行部分区域的值打印,这样打印出来的就不是带单元格类型的")
print(sheet1.col_values(3,2,5))
print(sheet1.row_values(0,0,6))


'''
xlwt 模块：只支持单独新建excel，不支持对已有的excle同步进行读写操作
         需要创建workbook对象，然后创建sheet对象，最后根据workbook对象的save方法保存为excel
'''
book=xlwt.Workbook()
sheet=book.add_sheet('firstsheet')
sheet.write(0,0,label='row0 ,column 0 value')
book.save('xlwt.xlsx')


'''
openpyxl 库 ：同时支持excel的读写，且对excel的的类型支持多，如xlsx xltx xlsm  lxtm 
          与wlrd的是打印单元格时，传入的值是单元格的坐标，即字符串 A5，而非2个数字

 基础概念   workbook  代表一个excel  ，  首先要load一个excle的路径，获取对象
      核心方法：cell( coordinate=None, row=None, column=None, value=None) 
         第一个参数是字符串的坐标如A5，第二个参数是行数字，第三个参数是列数字，第四个参数是写入的值  
          在这里他指的不是索引下标，而是真实的行号、列号
         
      代表一个单元格,传入【A5】进行获取，cell方法支持读写
         
       与wlrd  xlwt 库相比，它的API更具有面向对象的表述如：creat sheet ,copy worksheet,
      对单独一行或者一列的数据可以通过sheet对象切片来获取数据 sheet["C"]:打印一整列，返回一个生成器
       打印单独一个单元格可以通过 sheet["A2"].value   直接获取
       一些其他的方法也比较支持：如 冻结窗口、合并单元格
        
      最后调用excel对象的save方法保存excel ,可以重命名路径。如果没有调用save方法，文件内容是不会改变的。
'''
wb=openpyxl.load_workbook(filename=file)
sheetname=wb.sheetnames[0]
print(wb.sheetnames,"打印所有的sheet名字，这个库不支持通过索引获取sheet对象")
print(wb.active,"获取活跃的sheet")
print(wb.active["A2"].value)
print(wb.active.cell(row=2,column=2).value,"获取单元格值")
print(wb.active.cell("A2",value="A222222").value)
print(wb.active["A2"].value)
# print(wb.active.freeze_panes)
# print(wb.active.merge_cells(),"合并单元格")

'''

xlsxwriter :插入图片和图表,  最后通过关闭文件对象来保存操作
      通过sheet对象，调用insert_image  分别传入字符串的坐标， 和图片的地址 插入图片
       和insert_chart   传入字符串的坐标绝定左上角的起始位置，
'''
ef=xlsxwriter.Workbook(r'tubiao.xlsx')
worksheet=ef.add_worksheet()

worksheet.insert_image('B5','课程地址.png')
#worksheet.insert_chart()
#worksheet.insert_textbox
#worksheet.add_table
ef.close()



